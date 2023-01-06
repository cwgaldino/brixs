#! /usr/bin/env python3
# -*- coding: utf-8 -*-
"""Get data from xlsx (excel) files.

Author: Carlos Galdino
Creation date: 05/01/2023
"""

# %% ------------------------- Standard Imports --------------------------- %% #
from pathlib import Path
import types

# %% ------------------------- Special Imports ---------------------------- %% #
try:
    import openpyxl
except ModuleNotFoundError:
    pass

def xlsx(filepath, sheetname):
    filepath = Path(filepath)

    xlsx = openpyxl.load_workbook(str(filepath), data_only=True)
    sheet = xlsx[sheetname]
    sheet.filepath = filepath

    sheet.refresh = types.MethodType(refresh, sheet)

    sheet.is_merged = types.MethodType(is_merged, sheet)

    sheet.update_header = types.MethodType(update_header, sheet)
    sheet.check_header = types.MethodType(check_header, sheet)
    sheet.column_number_from_header = types.MethodType(column_number_from_header, sheet)

    sheet.get_column = types.MethodType(get_column, sheet)
    sheet.search_rows = types.MethodType(search_rows, sheet)
    sheet.get_dataset = types.MethodType(get_dataset, sheet)
    sheet.get_dataset_info = types.MethodType(get_dataset_info, sheet)


    sheet.update_header()

    return sheet

def refresh(self):
    filepath = Path(self.filepath)

    # self._parent.save(str(filepath.resolve()))
    sheetname = self.title
    sheet = initialize(filepath, sheetname)
    return sheet

def update_header(self, row=0):
    header = [None]*self.max_column
    for i, column in enumerate(self.iter_cols(min_row=row, max_row=row)):
        header[i] = column[0].value
    self.header = header

def check_header(self):
    if 'header' not in self.__dict__.keys():
        AttributeError('Header does not seem to be defined for this sheet.\nRun Sheet.update_header()')

def column_number_from_header(self, column):

    if type(column)==str:
        self.check_header()
        if column in self.header:
            column = self.header.index(column)+1
        else:
            raise ValueError(f'Cannot find {column} in header')
    return column

def get_column(self, column, min_row, max_row):
    # preallocation
    data = [None]*(max_row-min_row+1)

    # row
    assert max_row >= min_row, 'max_row must be bigger than min_row.'

    # column
    column = self.column_number_from_header(column)

    # get data
    for i, row in enumerate(self.iter_rows(min_row=min_row, max_row=max_row, min_col=column, max_col=column)):
        data[i] = row[0].value
    return data

def search_rows(self, string, column):
    column = self.column_number_from_header(column)
    for i, row in enumerate(self.iter_rows(min_col=column, max_col=column)):
        if string == row[0].value:
            return row[0].row

def is_merged(self, column, row):
    column = self.column_number_from_header(column)
    for ranges in self.merged_cells.ranges:
        if row >= ranges.min_row and row <= ranges.max_row:
            if column >= ranges.min_col and column <= ranges.max_col:
                return (ranges.min_col, ranges.max_col, ranges.min_row, ranges.max_row, )
    else:
        return False

def get_dataset_info(self, name, column):
    column_dataset = self.column_number_from_header('dataset')
    row = self.search_rows(name, column_dataset)

    if row is None:
        raise ValueError(f'Cannot find name: {name}')

    # get merged length
    range = self.is_merged(column=column_dataset, row=row)
    if range != False:
        info = self.get_column(column=column, min_row=range[2], max_row=range[3])
        if None in info:
            info = [x for x in info if x is not None]
        return info
    else:
        return self.get_column(column=column, min_row=row, max_row=row)

def get_dataset(self, name):
    column = self.column_number_from_header('dataset')
    row = self.search_rows(name, column)

    # get merged length
    range = self.is_merged(column, row)
    if range != False:
        print(range)
    else:
        print((column, column, row, row))

# # %%
# sheet = initialize('scan_book.xlsx', 'Sheet1')
#
# # %%
#
# sheet = sheet.refresh()
# sheet.get_dataset('momentum map, LV, (pi pi)')
#
#
# # %%
#
# # %% Testing
#
# # Define variable to load the dataframe
# print(xlsx.sheetnames)
#
# sheet = xlsx['Sheet1']
#
# sheet['B4'].__dir__()
# sheet['B5'].__dir__()
# sheet['B15'].value
# sheet['B15'].row
#
# for cell in sheet['B0:B31']:
#     # print(cell)
#     print(cell[0].value)
#
#
# sheet.__dir__()
# ranges.__dir__()
# def merged_size(start_row, start_column):
#
#     start_row    = 5
#     start_column = 1
#     type(sheet['B5'])
# sheet.__dir__()
