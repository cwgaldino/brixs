#! /usr/bin/env python3
# -*- coding: utf-8 -*-
"""Useful functions for everyday use ---> xlsx (excel) files"""

# %% ------------------------- Standard Imports --------------------------- %% #
from pathlib import Path
import types

# %% ============================== xlsx ================================== %% #
openpyxlok = False
try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font
    from openpyxl.utils import get_column_letter
    # from openpyxl.styles import Border, Side,  GradientFill, Alignment
    openpyxlok = True
except ModuleNotFoundError:
    pass

def new_xlsx(filepath, sheetname=None):
    """Open a new file

    Usage:
        # import
        from xlsx import xlsx, new_xlsx

        # create new xlsx file or open an existing file
        sheet = new_xlsx(filepath)  # new file
        sheet = xlsx(filepath)      # existing file

        # function names are self-explanatory
        sheet.set_row([1, 2, 3, 4], row=2, start_col=4)
        sheet.set_row_fill(color='ff0000', row=2, start_col=4, stop_col=8)
        sheet.set_row_font(row=2, bold=True, color='000000', start_col=4, stop_col=8)

        w = sheet.get_col_width(col=5)
        sheet.set_col_width(col=5, width=w*1.5)

        sheet.save()

    Args:
        filepath (Path or str): filepath for new xlsx file.
        sheetname (str, optional): name fo the first sheet (active sheet).
    
    Returns:
        sheet object
    """
    assert openpyxlok, 'new_xlsx() cannot open xlsx file\nError: python package `openpyxl` not found\nmaybe install it via ``pip install openpyxl``' 

    # initialize new xlsx
    xlsx = openpyxl.Workbook()

    # grab the first Sheet
    sheet = xlsx.active
    if sheetname is not None:
        sheet.title = str(sheetname)
    sheet.filepath = filepath

    # connect to methods
    _connect2methods(sheet)

    # save xlsx
    xlsx.save(str(filepath))

    return sheet

def _connect2methods(sheet):
    """Connect support methods to Sheet object.

    Args:
        sheet (sheet object): sheet to connect to support methods.

    Returns:
        None
    """
    sheet.save             = types.MethodType(_save, sheet)

    sheet.get_col_data     = types.MethodType(_get_col_data, sheet)
    sheet.get_row_data     = types.MethodType(_get_row_data, sheet)

    sheet.set_row          = types.MethodType(_set_row, sheet)
    sheet.set_row_fill     = types.MethodType(_set_row_fill, sheet)
    sheet.set_row_font     = types.MethodType(_set_row_font, sheet)
    sheet.set_col_width    = types.MethodType(_set_col_width, sheet)
    sheet.get_col_width    = types.MethodType(_get_col_width, sheet)

    # maybe obsolete
    # sheet.refresh          = types.MethodType(_refresh, sheet)
    # sheet.is_merged        = types.MethodType(_is_merged, sheet)
    # sheet.update_header    = types.MethodType(_update_header, sheet)
    # sheet.check_header     = types.MethodType(_check_header, sheet)
    # sheet.get_column       = types.MethodType(_get_column, sheet)
    # sheet.search_rows      = types.MethodType(_search_rows, sheet)
    # sheet.get_dataset      = types.MethodType(_get_dataset, sheet)
    # sheet.get_dataset_info = types.MethodType(_get_dataset_info, sheet)
    # sheet.column_number_from_header = types.MethodType(_column_number_from_header, sheet)
    return 

def xlsx(filepath, sheetname=None):
    """Open xlsx file

    Usage:
        # import
        from xlsx import xlsx, new_xlsx

        # create new xlsx file or open an existing file
        sheet = new_xlsx(filepath)  # new file
        sheet = xlsx(filepath)      # existing file

        # function names are self-explanatory
        sheet.set_row([1, 2, 3, 4], row=2, start_col=4)
        sheet.set_row_fill(color='ff0000', row=2, start_col=4, stop_col=8)
        sheet.set_row_font(row=2, bold=True, color='000000', start_col=4, stop_col=8)

        w = sheet.get_col_width(col=5)
        sheet.set_col_width(col=5, width=w*1.5)

        sheet.save()

    Args:
        filepath (Path, str): filepath to xlsx
        sheetname (str, optional): Sheet name. If None, the first sheet will be
            selected. Default is None.

    Return:
        sheet object.
    """
    assert openpyxlok, 'xlsx() cannot open xlsx file\nError: python package `openpyxl` not found\nmaybe install it via ``pip install openpyxl``' 

    filepath = Path(filepath)

    xlsx = openpyxl.load_workbook(str(filepath))#, data_only=True)

    if sheetname is None:
        sheet = xlsx.active
    else:
        sheet = xlsx[sheetname]
    sheet.filepath = filepath

    _connect2methods(sheet)

    return sheet

# %% ============================ Support methods ========================= %% #
def _save(self, filepath=None):
    """Save xlsx from sheet object
    
    Args:
        filepath (Path or str, optional): if None, it will search for filepath
            in Sheet.filepath
            
    Returns:
        None
    """
    if filepath is None:
        try:
            filepath = str(self.filepath)
        except AttributeError:
            raise AttributeError('sheet.filepath is not defined')
    
    self.parent.save(filepath)
    return

# %% ============================ Sheet manipulation ====================== %% #
def _get_col_data(self, col, start_row=1, stop_row=None):
    """get values from column

    Args:
        col (int): column number or letter
        start_row, stop_row (int): row to start and stop reading. Default is 
            start_row=1 and stop_row=None (max row with data)

    Returns:
        list
    """
    if isinstance(col, str) == False:
        col = get_column_letter(col)
    
    if stop_row is None:
        stop_row = self.max_row
    return [_[0].value for _ in self[f'{col}{start_row}:{col}{stop_row}']]

def _get_row_data(self, row, start_col=1, stop_col=None):
    """get values from row

    Args:
        row (int): row number
        start_col, stop_col (int): column to start and stop reading. Default is 
            start_col=1 and stop_col=None (max column with data)

    Returns:
        list
    """    
    assert row > 0, 'row cannot 0 or negative'
    if stop_col is None:
        stop_col = self.max_column
    
    if isinstance(start_col, str) == False:
        start_col = get_column_letter(start_col)
    if isinstance(stop_col, str) == False:
        stop_col = get_column_letter(stop_col)
        
    return [_.value for _ in self[f'{start_col}{row}:{stop_col}{row}'][0]]


def _set_row(self, values, row, start_col=1):
    """Write values in a row.

    Args:
        values (list): list with values
        row (int): row number
        start_col (int, optional): column to start placing values. Default is 1

    Returns:
        None
    """
    for col, val in enumerate(values, start=0):
        self.cell(row=row, column=start_col+col).value = val   

def _set_row_fill(self, color, row, start_col=1, stop_col=None):
    """Write values in a row.

    Args:
        color (str): HEX value of a color
        row (int): row number
        start_col (int, optional): column to start filling. Default is 1
        stop_col (int, optional): column to stop filling. If None, the max
            number of used columns will be used. Default is None.

    Returns:
        None
    """
    for col  in range(start_col, stop_col+1):
        self.cell(row=row, column=col).fill = PatternFill("solid", fgColor=color)

def _set_row_font(self, row, bold=False, color='000000', start_col=1, stop_col=None):
    """Write values in a row.

    Args:
        row (int): row number
        bold (bool): Bold if True
        color (str): HEX value of a color
        start_col (int, optional): column to start filling. Default is 1
        stop_col (int, optional): column to stop filling. If None, the max
            number of used columns will be used. Default is None.

    Returns:
        None
    """
    for col  in range(start_col, stop_col+1):
        self.cell(row=row, column=col).font = Font(b=bold, color=color)

def _set_col_width(self, col, width):
    """Set column width
    
    Args:
        col (str or int): columns number or letter
        width (int): width value
        
    Returns:
        None
    """
    if isinstance(col, str) == False:
        col = get_column_letter(col)
    self.column_dimensions[col].width = width

def _get_col_width(self, col):
    """Set column width
    
    Args:
        col (str or int): columns number or letter
        width (int): width value
        
    Returns:
        None
    """
    if isinstance(col, str) == False:
        col = get_column_letter(col)
    return self.column_dimensions[col].width


# %% =============================== Obsolete ============================= %% #
# def _refresh(self):
#     filepath = Path(self.filepath)

#     # self._parent.save(str(filepath.resolve()))
#     sheetname = self.title
#     sheet     = xlsx(filepath, sheetname)
#     return sheet

# def _update_header(self, row=0):
#     header = [None]*self.max_column
#     for i, column in enumerate(self.iter_cols(min_row=row, max_row=row)):
#         header[i] = column[0].value
#     self.header = header

# def _check_header(self):
#     if 'header' not in self.__dict__.keys():
#         AttributeError('Header does not seem to be defined for this sheet.\nRun Sheet.update_header()')

# def _column_number_from_header(self, column):

#     if type(column)==str:
#         self.check_header()
#         if column in self.header:
#             column = self.header.index(column)+1
#         else:
#             raise ValueError(f'Cannot find {column} in header')
#     return column

# def _get_column(self, column, min_row, max_row):
#     # preallocation
#     data = [None]*(max_row-min_row+1)

#     # row
#     assert max_row >= min_row, 'max_row must be bigger than min_row.'

#     # column
#     column = self.column_number_from_header(column)

#     # get data
#     for i, row in enumerate(self.iter_rows(min_row=min_row, max_row=max_row, min_col=column, max_col=column)):
#         data[i] = row[0].value
#     return data

# def _search_rows(self, string, column):
#     column = self.column_number_from_header(column)
#     for i, row in enumerate(self.iter_rows(min_col=column, max_col=column)):
#         if string == row[0].value:
#             return row[0].row

# def _is_merged(self, column, row):
#     column = self.column_number_from_header(column)
#     for ranges in self.merged_cells.ranges:
#         if row >= ranges.min_row and row <= ranges.max_row:
#             if column >= ranges.min_col and column <= ranges.max_col:
#                 return (ranges.min_col, ranges.max_col, ranges.min_row, ranges.max_row, )
#     else:
#         return False

# def _get_dataset_info(self, name, column):
#     column_dataset = self.column_number_from_header('dataset')
#     row = self.search_rows(name, column_dataset)

#     if row is None:
#         raise ValueError(f'Cannot find name: {name}')

#     # get merged length
#     range = self.is_merged(column=column_dataset, row=row)
#     if range != False:
#         info = self.get_column(column=column, min_row=range[2], max_row=range[3])
#         if None in info:
#             info = [x for x in info if x is not None]
#         return info
#     else:
#         return self.get_column(column=column, min_row=row, max_row=row)

# def _get_dataset(self, name):
#     column = self.column_number_from_header('dataset')
#     row = self.search_rows(name, column)

#     # get merged length
#     range = self.is_merged(column, row)
#     if range != False:
#         print(range)
#     else:
#         print((column, column, row, row))


